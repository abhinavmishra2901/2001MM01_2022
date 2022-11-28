# Project 2 CS384 -     Octant Batch Processing and Merging of Assignment Tut01-05.
# By Abhinav Mishra - 2001MM01 and Hardik Tiwari - 2001MM15

# Libraries

from datetime import datetime

import os
import openpyxl_dictreader
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side, BORDER_THIN
import math
import streamlit as st
import shutil

os.system("cls")
start_time = datetime.now()

# Help

# Read all the excel files in a batch format from the input/ folder. Only xlsx to be allowed
# Save all the excel files in a the output/ folder. Only xlsx to be allowed
# output filename = input_filename[_octant_analysis_mod_5000].xlsx , ie, append _octant_analysis_mod_5000 to the original filename.

# Code

# Function - 1 to count the octant ranks


def octant_rank_count(count):
    # Overall Rank
    overall_rank = []
    # Creating a copy of the attribute list
    count_copy = count.copy()
    # Reversing the copy of the attribute list
    count_copy.sort(reverse=True)
    # Checking and appending the index of the sorted list's elements in the original list
    for i in count_copy:
        overall_rank.append(count.index(i))

    # Storing the ranks to an octant_rank_list by mapping the indexes
    octant_rank_list = [0]*8
    rank = 1
    for j in range(8):
        check = overall_rank[j]
        octant_rank_list[check] = rank
        rank += 1

    # Condition if two ranks are equal
    for j in range(8):
        if octant_rank_list[j] == 0:
            for x in range(1, 9):
                if x not in octant_rank_list:
                    octant_rank_list[j] = x

    return octant_rank_list

# Function - 2 to write the octant range names


def octant_range_names(sheet, inputfile, mod):
    octant_name_id_mapping = {"1": "Internal outward interaction", "-1": "External outward interaction", "2": "External Ejection",
                              "-2": "Internal Ejection", "3": "External inward interaction", "-3": "Internal inward interaction", "4": "Internal sweep", "-4": "External sweep"}
    # Declaring the lists to store the values
    time = []
    u = []
    v = []
    w = []
    u_prime = []
    v_prime = []
    w_prime = []
    octant = []

    # Opening the octant_input.xlsx file in read mode
    try:
        reader = openpyxl_dictreader.DictReader(
            inputfile, "Sheet1")

        # Storing the values of each key in the corresponding lists
        for row in reader:
            time.append(float(row['T']))
            u.append(float(row['U']))
            v.append(float(row['V']))
            w.append(float(row['W']))

        # Calculating the average of U, V, W
        u_avg = round(sum(u)/len(u), 3)
        v_avg = round(sum(v)/len(v), 3)
        w_avg = round(sum(w)/len(w), 3)

        # Data Preprocessing - Calculating the difference between the velocities and their respective average values and storing in the respective lists.
        for u_value in u:
            u_prime.append(round(u_value-u_avg, 3))
        for v_value in v:
            v_prime.append(round(v_value-v_avg, 3))
        for w_value in w:
            w_prime.append(round(w_value-w_avg, 3))

    # FileNotFound Error
    except FileNotFoundError:
        st.markdown("Input File not available!")
        st.stop()

    # Other errors if any
    except:
        st.markdown("Some error occured while reading the input file")
        st.stop()

    # Declaring List to store the count of each Octant ID
    count = [0]*8

    # Tagging the octants by help of the video provided in the assignment
    for i in range(0, len(time)):
        if (u_prime[i] >= 0 and v_prime[i] >= 0):
            if w_prime[i] >= 0:
                octant.append(1)
                count[0] += 1
            else:
                octant.append(-1)
                count[1] += 1
        if (u_prime[i] < 0 and v_prime[i] >= 0):
            if w_prime[i] >= 0:
                octant.append(2)
                count[2] += 1
            else:
                octant.append(-2)
                count[3] += 1
        if (u_prime[i] < 0 and v_prime[i] < 0):
            if w_prime[i] >= 0:
                octant.append(3)
                count[4] += 1
            else:
                octant.append(-3)
                count[5] += 1
        if (u_prime[i] >= 0 and v_prime[i] < 0):
            if w_prime[i] >= 0:
                octant.append(4)
                count[6] += 1
            else:
                octant.append(-4)
                count[7] += 1

    # Defining Ranges with help of mod
    range1 = []

    # Defining a list to store the ranks
    rank_list = []
    rank_list.append(octant_rank_count(count))
    rank_list.append([""]*8)

    # With each step in the loop, if the condition is satisfied we increase the range count by 1 and also append a blank space in label for convenience in later steps
    for x in range(0, len(time), mod):
        if x == 0:
            range1.append(".0000-{}".format(mod-1))
        elif x+mod > len(time):
            range1.append("{}-{}".format(x, len(time)-1))
        else:
            range1.append("{}-{}".format(x, x+mod-1))

    # Declaring Lists to store the count of each octant ID in the given mod
    mod_c1 = []
    mod_cm1 = []
    mod_c2 = []
    mod_cm2 = []
    mod_c3 = []
    mod_cm3 = []
    mod_c4 = []
    mod_cm4 = []
    # Here cmi refers to -ith octant

    # Counting the octant values in each mod using a loop method for a particular range in the octant list
    z = 0
    y = mod
    while y-mod < len(time):
        mod_c1.append(octant[z:y].count(1))
        mod_cm1.append(octant[z:y].count(-1))
        mod_c2.append(octant[z:y].count(2))
        mod_cm2.append(octant[z:y].count(-2))
        mod_c3.append(octant[z:y].count(3))
        mod_cm3.append(octant[z:y].count(-3))
        mod_c4.append(octant[z:y].count(4))
        mod_cm4.append(octant[z:y].count(-4))
        z = y
        y = y+mod

    # Storing the ranks of the mod ranges in the rank_list variable. Here the elements are initially row-wise and I am changing it to column-wise to print later.
    for i in range(len(mod_c1)):
        mod_list = []
        mod_list.append(mod_c1[i])
        mod_list.append(mod_cm1[i])
        mod_list.append(mod_c2[i])
        mod_list.append(mod_cm2[i])
        mod_list.append(mod_c3[i])
        mod_list.append(mod_cm3[i])
        mod_list.append(mod_c4[i])
        mod_list.append(mod_cm4[i])
        rank_list.append(octant_rank_count(mod_list))

    # Creating column-wise list to store the ranks in the designated order
    rank1 = []
    rank2 = []
    rank3 = []
    rank4 = []
    rank5 = []
    rank6 = []
    rank7 = []
    rank8 = []

    # Creating lists to store the index of the first rank, the first rank octant id and first rank octant name
    first_rank_index = []
    first_rank = []
    first_rank_name = []
    octant_id_list = ["1", "-1", "2", "-2", "3", "-3", "4", "-4"]

    # Running a loop to map the index of rank 1 to the corresponding lists and then map the octant id to the octant name by using the octant_name_id_mapping dictionary
    for i in range(len(rank_list)):
        if 1 in rank_list[i]:
            first_rank_index.append(rank_list[i].index(1))
    for i in range(len(first_rank_index)):
        first_rank.append(octant_id_list[first_rank_index[i]])
    for i in range(len(first_rank)):
        first_rank_name.append(octant_name_id_mapping[str(first_rank[i])])
    for i in range(len(rank_list)):
        rank1.append(rank_list[i][0])
        rank2.append(rank_list[i][1])
        rank3.append(rank_list[i][2])
        rank4.append(rank_list[i][3])
        rank5.append(rank_list[i][4])
        rank6.append(rank_list[i][5])
        rank7.append(rank_list[i][6])
        rank8.append(rank_list[i][7])

    # Extending the 3 columns to store the rank1 mod values count
    rank7.extend(("", "", "", "Octant ID"))
    rank7.extend(octant_id_list)
    rank8.extend(("", "", "", "Octant Name"))
    octant_name_list = []
    for i in range(8):
        octant_name_list.append(octant_name_id_mapping[str(octant_id_list[i])])
    rank8.extend(octant_name_list)
    first_rank.extend(("", "", "", "Count of Rank 1 Mod Values"))
    rank1_mod_values = []

    # Here we are slicing the rank1 list to exclude the overall rank1 count
    rank1_mod_values.append(rank1[2:].count(1))
    rank1_mod_values.append(rank2[2:].count(1))
    rank1_mod_values.append(rank3[2:].count(1))
    rank1_mod_values.append(rank4[2:].count(1))
    rank1_mod_values.append(rank5[2:].count(1))
    rank1_mod_values.append(rank6[2:].count(1))
    rank1_mod_values.append(rank7[2:].count(1))
    rank1_mod_values.append(rank8[2:].count(1))
    first_rank.extend(rank1_mod_values)

    # Appending the remaining length of the lists with a blank string for convenience in later steps
    for x in range(int(len(time)/mod)+2, len(time)):
        mod_c1.append("")
        mod_cm1.append("")
        mod_c2.append("")
        mod_cm2.append("")
        mod_c3.append("")
        mod_cm3.append("")
        mod_c4.append("")
        mod_cm4.append("")
        range1.append("")
        rank1.append("")
        rank2.append("")
        rank3.append("")
        rank4.append("")
        rank5.append("")
        rank6.append("")
        rank7.append("")
        rank8.append("")
        first_rank.append("")
        first_rank_name.append("")

    try:
        # Defining thin_border
        bd = Side(border_style='thin')
        thin_border = Border(left=bd, top=bd, right=bd, bottom=bd)

        # Header line1
        header_line1 = ["", "", "", "", "", "", "", "", "",
                        "", "", " ", " ", "Overall Octant Count", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]
        # Header line2
        header_line2 = ["Time", "U", "V", "W", "U Avg", "V Avg", "W Avg", "U'=U-U avg", "V'=V-V avg",
                        "W'=W-W avg", "Octant", " ", " ", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]

        # Loop to print the header line1
        for i in range(1, 27):
            # The Cell Address. Converting integer to the corresponding character by ascii conversion
            cell = chr(i+64)+'1'
            sheet[cell] = header_line1[i-1]
        for i in range(1, 7):
            # The Cell Address for AA to AF. Converting integer to the corresponding character by ascii conversion
            cell = 'A'+chr(i+64)+'1'
            sheet[cell] = header_line1[i+25]

        # Loop to print the header line2
        for i in range(1, 27):
            # The Cell Address. Converting integer to the corresponding character by ascii conversion
            cell = chr(i+64)+'2'
            sheet[cell] = header_line2[i-1]
        for i in range(1, 7):
            # The Cell Address for AA to AF. Converting integer to the corresponding character by ascii conversion
            cell = 'A'+chr(i+64)+'2'
            sheet[cell] = header_line2[i+25]

        # Writing the first two line separately due to difference in the data length
        output_row_1 = [time[0], u[0], v[0], w[0], u_avg, v_avg, w_avg, u_prime[0], v_prime[0],
                        w_prime[0], octant[0], "", "Octant ID", "1", "-1", "2", "-2", "3", "-3", "4", "-4", "Rank of 1", "Rank of -1", "Rank of 2", "Rank of -2", "Rank of 3", "Rank of -3", "Rank of 4", "Rank of -4", "Rank 1 Octant ID", "Rank 1 Octant Name"]
        output_row_2 = [time[1], u[1], v[1], w[1], "", "", "", u_prime[1], v_prime[1],
                        w_prime[1], octant[1], "Mod {}".format(mod), "Overall Count", count[0], count[1], count[2], count[3], count[4], count[5], count[6], count[7], rank1[0], rank2[0], rank3[0], rank4[0], rank5[0], rank6[0], rank7[0], rank8[0], first_rank[0], first_rank_name[0]]

        # Loop to print the first output line into the cells
        for i in range(1, 27):
            # Condition to leave column L empty
            if i == 12:
                continue
            # The Cell Address. Converting integer to the corresponding character by ascii conversion
            cell = chr(i+64)+'3'
            if i > 12:
                sheet[cell] = output_row_1[i-2]
            else:
                sheet[cell] = output_row_1[i-1]

        for i in range(1, 7):
            # The Cell Address for AA to AF. Converting integer to the corresponding character by ascii conversion
            cell = 'A'+chr(i+64)+'3'
            sheet[cell] = output_row_1[i+24]

        # Loop to print the second output line into the cells
        for i in range(1, 27):
            # Condition to leave column L empty
            if i == 12:
                continue
            # The Cell Address. Converting integer to the corresponding character by ascii conversion
            cell = chr(i+64)+'4'
            if i > 12:
                sheet[cell] = output_row_2[i-2]
            else:
                sheet[cell] = output_row_2[i-1]

        for i in range(1, 7):
            # The Cell Address for AA to AF. Converting integer to the corresponding character by ascii conversion
            cell = 'A'+chr(i+64)+'4'
            sheet[cell] = output_row_2[i+24]

        # Declaring a list to store the output of remaining lines/rows
        output_row = []
        for i in range(2, len(time)):
            output_row.append([time[i], u[i], v[i], w[i], " ", " ", " ", u_prime[i], v_prime[i], w_prime[i], octant[i], " ", range1[i-2],
                               mod_c1[i-2], mod_cm1[i-2], mod_c2[i-2], mod_cm2[i-2], mod_c3[i-2], mod_cm3[i-2], mod_c4[i-2], mod_cm4[i-2], rank1[i], rank2[i], rank3[i], rank4[i], rank5[i], rank6[i], rank7[i], rank8[i], first_rank[i-1], first_rank_name[i-1]])

        # Writing the remaining values to the output file
        # Here i is the range of columns and j is the range of rows. By combinations of characters we are storing the data to the corresponding cells.
        for i in range(1, 27):
            # Condition to leave column L empty
            if i == 12:
                continue
            if i > 12:
                for j in range(5, len(time)+2):
                    cell = chr(i+64)+str(j)
                    sheet[cell] = output_row[j-5][i-2]

            else:
                for j in range(5, len(time)+2):
                    cell = chr(i+64)+str(j)
                    sheet[cell] = output_row[j-5][i-1]

        # The Cell Address for AA to AF. Converting integer to the corresponding character by ascii conversion
        for i in range(1, 7):
            for j in range(5, len(time)+2):
                cell = 'A'+chr(i+64)+str(j)
                sheet[cell] = output_row[j-5][i+24]

        # Highlighting the Rank-1 Values
        for i in range(23, 27):
            for j in range(4, 5+int(len(time)/mod)+1):
                cell = chr(i+64)+str(j)
                sheet[cell].border = thin_border
                if sheet[cell].value == 1:
                    highlight = PatternFill(
                        start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    sheet[cell].fill = highlight
        for i in range(1, 5):
            for j in range(4, 5+int(len(time)/mod)+1):
                cell = 'A'+chr(i+64)+str(j)
                sheet[cell].border = thin_border
                if sheet[cell].value == 1:
                    highlight = PatternFill(
                        start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    sheet[cell].fill = highlight

        # Adding Borders to the required cells
        for i in range(14, 27):
            for j in range(3, 3+int(len(time)/mod)+3):
                cell = chr(i+64)+str(j)
                sheet[cell].border = thin_border
        for i in range(1, 7):
            for j in range(3, 3+int(len(time)/mod)+3):
                cell = 'A'+chr(i+64)+str(j)
                sheet[cell].border = thin_border
        for i in range(3, 6):
            for j in range(3+int(len(time)/mod)+5, 3+int(len(time)/mod)+4+11):
                cell = 'A'+chr(i+64)+str(j)
                sheet[cell].border = thin_border
    except:
        print("Something went wrong while writing to octant_output.csv")
        st.stop()

# Function - 3 to write the transition count


def octant_transition_count(sheet, inputfile, mod):

    # Declaring the lists to store the values
    time = []
    u = []
    v = []
    w = []
    u_prime = []
    v_prime = []
    w_prime = []
    octant = []

    # Opening the input_octant_transition_identify.xlsx file in read mode
    try:
        reader = openpyxl_dictreader.DictReader(
            inputfile, "Sheet1")

        # Storing the values of each key in the corresponding lists
        for row in reader:
            time.append(float(row['T']))
            u.append(float(row['U']))
            v.append(float(row['V']))
            w.append(float(row['W']))

        # Calculating the average of U, V, W
        u_avg = round(sum(u)/len(u), 3)
        v_avg = round(sum(v)/len(v), 3)
        w_avg = round(sum(w)/len(w), 3)

        # Data Preprocessing - Calculating the difference between the velocities and their respective average values and storing in the respective lists.
        for u_value in u:
            u_prime.append(round(u_value-u_avg, 3))
        for v_value in v:
            v_prime.append(round(v_value-v_avg, 3))
        for w_value in w:
            w_prime.append(round(w_value-w_avg, 3))
    # FileNotFound Error
    except FileNotFoundError:
        st.markdown("Input File not found!")
        st.stop()
    # Other errors if any
    except:
        st.markdown("Some error occured while reading the input file")
        st.stop()

    # Declaring List to store the count of each Octant ID
    count = [0]*8

    # Tagging the octants by help of the video provided in the assignment
    for i in range(0, len(time)):
        if (u_prime[i] >= 0 and v_prime[i] >= 0):
            if w_prime[i] >= 0:
                octant.append(1)
                count[0] += 1
            else:
                octant.append(-1)
                count[1] += 1
        if (u_prime[i] < 0 and v_prime[i] >= 0):
            if w_prime[i] >= 0:
                octant.append(2)
                count[2] += 1
            else:
                octant.append(-2)
                count[3] += 1
        if (u_prime[i] < 0 and v_prime[i] < 0):
            if w_prime[i] >= 0:
                octant.append(3)
                count[4] += 1
            else:
                octant.append(-3)
                count[5] += 1
        if (u_prime[i] >= 0 and v_prime[i] < 0):
            if w_prime[i] >= 0:
                octant.append(4)
                count[6] += 1
            else:
                octant.append(-4)
                count[7] += 1

    # Defining Ranges with help of mod
    range1 = []

    # Also defining a list called label to store some label texts for column L in the excel sheet
    label = [" "]
    range_count = 0

    # With each step in the loop, if the condition is satisfied we increase the range count by 1 and also append a blank space in label for convenience in later steps
    for x in range(0, len(time), mod):
        if x == 0:
            range1.append(".0000-{}".format(mod-1))
            range_count = range_count+1
        elif x+mod > len(time):
            range1.append("{}-{}".format(x, len(time)-1))
            range_count = range_count+1
        else:
            range1.append("{}-{}".format(x, x+mod-1))
            range_count = range_count+1

    # Extending range1 and label list
    range1.extend(("Overall Transition Count", " ", "Octant #",
                   "+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"))
    label.extend(("", "", "", "From", "", "", "", "", "", "", ""))

    # Declaring Lists to store the count of each octant ID in the given mod
    mod_c1 = []
    mod_cm1 = []
    mod_c2 = []
    mod_cm2 = []
    mod_c3 = []
    mod_cm3 = []
    mod_c4 = []
    mod_cm4 = []
    # Here cmi refers to -ith octant

    # Overall Transition Count
    overall_transition_list = []

    # Counting the octant values in each mod using a loop method for a particular range in the octant list
    z = 0
    y = mod
    while y-mod < len(time):
        mod_c1.append(octant[z:y].count(1))
        mod_cm1.append(octant[z:y].count(-1))
        mod_c2.append(octant[z:y].count(2))
        mod_cm2.append(octant[z:y].count(-2))
        mod_c3.append(octant[z:y].count(3))
        mod_cm3.append(octant[z:y].count(-3))
        mod_c4.append(octant[z:y].count(4))
        mod_cm4.append(octant[z:y].count(-4))
        z = y
        y = y+mod

    # In this step we also extend the lists to print the column headers for overall transition count
    mod_c1.extend((" ", "To", "+1"))
    mod_cm1.extend((" ", " ", "-1"))
    mod_c2.extend((" ", " ", "+2"))
    mod_cm2.extend((" ", " ",  "-2"))
    mod_c3.extend((" ", " ", "+3"))
    mod_cm3.extend((" ", " ",  "-3"))
    mod_c4.extend((" ", " ", "+4"))
    mod_cm4.extend((" ", " ",  "-4"))

    # A simple 2D loop to check the transitions. I am using two loops and then adding the count to overall_transition list
    highest_highlights = []
    for j in range(8):
        octant_list = [1, -1, 2, -2, 3, -3, 4, -4]
        overall_transition = [0]*8
        for i in range(len(time)-1):
            if octant[i+1] == octant_list[j]:
                if octant[i] == 1:
                    overall_transition[0] += 1
                elif octant[i] == -1:
                    overall_transition[1] += 1
                elif octant[i] == 2:
                    overall_transition[2] += 1
                elif octant[i] == -2:
                    overall_transition[3] += 1
                elif octant[i] == 3:
                    overall_transition[4] += 1
                elif octant[i] == -3:
                    overall_transition[5] += 1
                elif octant[i] == 4:
                    overall_transition[6] += 1
                elif octant[i] == -4:
                    overall_transition[7] += 1

        # overall_transition_list is the list to store the 8 lists we obtained in the above loop
        overall_transition_list.append(overall_transition)

    highest_highlight = []
    highlight_cell_lists = []
    for i in range(len(overall_transition_list)):
        highlight_cell_list = []
        for otl in overall_transition_list:
            highlight_cell_list.append(otl[i])
        highlight_cell_lists.append(highlight_cell_list)
        highest_highlight.append(max(highlight_cell_list))
    highest_highlights.append(highest_highlight)

    # To print the data to the excel file, we append the overall_transition_list's values to the respective columns using a loop
    for i in range(8):
        mod_c1.append(overall_transition_list[0][i])
        mod_cm1.append(overall_transition_list[1][i])
        mod_c2.append(overall_transition_list[2][i])
        mod_cm2.append(overall_transition_list[3][i])
        mod_c3.append(overall_transition_list[4][i])
        mod_cm3.append(overall_transition_list[5][i])
        mod_c4.append(overall_transition_list[6][i])
        mod_cm4.append(overall_transition_list[7][i])
    overall_transition_list.clear()

    # Mod Transition Counts
    # We start with a loop iterating the upper round off of 30000/mod
    z = 0
    y = mod
    for x in range(math.ceil(len(time)/mod)):
        # A special if condition, to handle the last case
        if y >= len(time):
            y = len(octant)-1

        # Extending the lists to label the lists for each iteration
        range1.extend(("", "", " ", "Mod Transition Count",
                       range1[x], "Octant #", "+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"))
        label.extend(("", "", "", "", "", "", "From",
                      "", "", "", "", "", "", ""))
        mod_c1.extend(("", " ", " ", " ", "To", "+1"))
        mod_cm1.extend(("", " ", " ", " ", " ", "-1"))
        mod_c2.extend(("", " ", " ", " ", " ", "+2"))
        mod_cm2.extend(("", " ", " ", " ", " ", "-2"))
        mod_c3.extend(("", " ", " ", " ", " ", "+3"))
        mod_cm3.extend(("", " ", " ", " ", " ", "-3"))
        mod_c4.extend(("", " ", " ", " ", " ", "+4"))
        mod_cm4.extend(("", " ", " ", " ", " ", "-4"))

        # A simple 2D loop to check the transitions. I am using two loops and then adding the count to overall_transition list
        for j in range(8):
            octant_list = [1, -1, 2, -2, 3, -3, 4, -4]
            overall_transition = [0]*8
            for i in range(z, y):
                if octant[i+1] == octant_list[j]:
                    if octant[i] == 1:
                        overall_transition[0] += 1
                    elif octant[i] == -1:
                        overall_transition[1] += 1
                    elif octant[i] == 2:
                        overall_transition[2] += 1
                    elif octant[i] == -2:
                        overall_transition[3] += 1
                    elif octant[i] == 3:
                        overall_transition[4] += 1
                    elif octant[i] == -3:
                        overall_transition[5] += 1
                    elif octant[i] == 4:
                        overall_transition[6] += 1
                    elif octant[i] == -4:
                        overall_transition[7] += 1

        # overall_transition_list is the list to store the 8 lists we obtained in the above loop
            overall_transition_list.append(overall_transition)

        highest_highlight = []
        highlight_cell_lists = []
        for i in range(len(overall_transition_list)):
            highlight_cell_list = []
            for otl in overall_transition_list:
                highlight_cell_list.append(otl[i])
            highlight_cell_lists.append(highlight_cell_list)
            highest_highlight.append(max(highlight_cell_list))
        highest_highlights.append(highest_highlight)

    # To print the data to the excel file, we append the overall_transition_list's values to the respective columns using a loop
        for i in range(8):
            mod_c1.append(overall_transition_list[0][i])
            mod_cm1.append(overall_transition_list[1][i])
            mod_c2.append(overall_transition_list[2][i])
            mod_cm2.append(overall_transition_list[3][i])
            mod_c3.append(overall_transition_list[4][i])
            mod_cm3.append(overall_transition_list[5][i])
            mod_c4.append(overall_transition_list[6][i])
            mod_cm4.append(overall_transition_list[7][i])
        z = y
        y = y+mod
        # Clearing the lists for the iteration of next range
        overall_transition_list.clear()

    # Appending the remaining length of the lists with a blank string for convenience in later steps
    for x in range(int(len(time)/mod)+2, len(time)):
        mod_c1.append("")
        mod_cm1.append("")
        mod_c2.append("")
        mod_cm2.append("")
        mod_c3.append("")
        mod_cm3.append("")
        mod_c4.append("")
        mod_cm4.append("")

    # Appending the remaining length of the lists with a blank string for convenience in later steps
    for x in range(range_count+1, len(time)):
        range1.append("")
        label.append("")

    try:
        # Defining thin_border
        bd = Side(border_style='thin')
        thin_border = Border(left=bd, top=bd, right=bd, bottom=bd)

        # Declaring a list to store the output of lines/rows
        output_row = []
        for i in range(0, len(time)):
            j = math.ceil(len(time)/mod)
            output_row.append([label[i+1], range1[i+j], mod_c1[i+j], mod_cm1[i+j], mod_c2[i+j],
                               mod_cm2[i+j], mod_c3[i+j], mod_cm3[i+j], mod_c4[i+j], mod_cm4[i+j]])

        # Writing the remaining values to the output file
        # Here i is the range of columns and j is the range of rows. By combinations of characters we are storing the data to the corresponding cells.
        for i in range(8, 18):
            for j in range(1, len(time)):
                cell = 'A'+chr(i+64)+str(j)
                sheet[cell] = output_row[j-1][i-8]

        # Applying borders to the required cells
        for i in range(9, 18):
            for j in range(3, 12):
                cell = 'A'+chr(i+64)+str(j)
                sheet[cell].border = thin_border

        start_row = 17
        end_row = 26
        for i in range(int(len(time)/mod)+1):
            for j in range(9, 18):
                for k in range(start_row, end_row):
                    cell = 'A'+chr(j+64)+str(k)
                    sheet[cell].border = thin_border
            start_row += 14
            end_row += 14
        # Loop for Highlighting the Highest Transition Values
        start_row = 4
        end_row = 12
        check = 0
        while end_row < len(time):
            for i in range(start_row, end_row):
                row_check = 0
                for j in range(len(highest_highlights)):
                    for k in range(10, 18):
                        cell = 'A'+chr(k+64)+str(i)
                        if row_check == 0:
                            if sheet[cell].value == highest_highlights[j][check]:
                                highlight = PatternFill(
                                    start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                                sheet[cell].fill = highlight
                                check += 1
                                row_check = 1
                                if check > 7:
                                    check = 0
            start_row += 14
            end_row += 14
    except:
        print("Something went wrong while writing to octant_output.csv")
        st.stop()

# Function - 4 to write the longest subsequence count with range


def octant_longest_subsequence_count_with_range(sheet, inputfile):

    # Declaring the lists to store the values
    time = []
    u = []
    v = []
    w = []
    u_prime = []
    v_prime = []
    w_prime = []
    octant = []

    # Opening the input_octant_transition_identify.xlsx file in read mode
    try:
        reader = openpyxl_dictreader.DictReader(
            inputfile, "Sheet1")

        # Storing the values of each key in the corresponding lists
        for row in reader:
            time.append(float(row['T']))
            u.append(float(row['U']))
            v.append(float(row['V']))
            w.append(float(row['W']))

        # Calculating the average of U, V, W
        u_avg = round(round(sum(u)/len(u), 3), 3)
        v_avg = round(round(sum(v)/len(v), 3), 3)
        w_avg = round(round(sum(w)/len(w), 3), 3)

        # Data Preprocessing - Calculating the difference between the velocities and their respective average values and storing in the respective lists.
        for u_value in u:
            u_prime.append(u_value-u_avg)
        for v_value in v:
            v_prime.append(v_value-v_avg)
        for w_value in w:
            w_prime.append(w_value-w_avg)

    # FileNotFound Error
    except FileNotFoundError:
        st.markdown("Input File not found!")
        st.stop()

    # Other errors if any
    except:
        st.markdown("Some error occured while reading the input file")
        st.stop()

    # Declaring List to store the count of each Octant ID
    count = [0]*8
    long_sub_len = []
    count_sub_len = []

    # Declaring Lists to store the count of ranges
    long_sub_len_range = []
    count_sub_len_range = []

    # Tagging the octants by help of the video provided in the assignment
    for i in range(0, len(time)):
        if (u_prime[i] >= 0 and v_prime[i] >= 0):
            if w_prime[i] >= 0:
                octant.append("+1")
                count[0] += 1
            else:
                octant.append("-1")
                count[1] += 1
        if (u_prime[i] < 0 and v_prime[i] >= 0):
            if w_prime[i] >= 0:
                octant.append("+2")
                count[2] += 1
            else:
                octant.append("-2")
                count[3] += 1
        if (u_prime[i] < 0 and v_prime[i] < 0):
            if w_prime[i] >= 0:
                octant.append("+3")
                count[4] += 1
            else:
                octant.append("-3")
                count[5] += 1
        if (u_prime[i] >= 0 and v_prime[i] < 0):
            if w_prime[i] >= 0:
                octant.append("+4")
                count[6] += 1
            else:
                octant.append("-4")
                count[7] += 1

    try:
        # Defining thin_border
        bd = Side(border_style='thin')
        thin_border = Border(left=bd, top=bd, right=bd, bottom=bd)

        # Header Labels
        sheet['AS1'] = "Longest Subsequence Length"
        sheet['AW1'] = "Longest Subsequence Length with Range"

        # Header line
        header_line = ["Octant ##", "Longest Subsequence Length", "Count",
                       "", "Octant ###", "Longest Subsequence Length", "Count"]

        # Loop to print the header line
        for i in range(19, 26):
            # The Cell Address. Converting integer to the corresponding character by ascii conversion
            cell = 'A'+chr(i+64)+'3'
            sheet[cell] = header_line[i-19]

        # Declaring a list to store the output of lines/rows
        output_row = []

        # Declaring a list to store the octants
        octant_ids = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
        octant_ids2 = []  # Declaring octant_ids2 list for a different column with "Time" label

        # Loop to check the maximum consecutive subsequence as well as the count or repetition of the subsequence
        # Iterating through octant_ids
        for j in octant_ids:
            octant_ids2.append(j)
            # Declaring variables to show the current maximum length of the subsequence and the previous maximum length of the subsequence
            previous_count = 0
            max_count = 0
            # Iterating through the octant list and checking the maximum length of a subsequence
            for i in range(len(time)):
                if octant[i] == j:
                    previous_count += 1
                else:
                    if previous_count > max_count:
                        max_count = previous_count
                    previous_count = 0

            # Next I am declaring variables to show the count of repetition of the subsequence in the octant list
            range_count = 0
            count_check = 0
            # Also declaring variables to store the from, to and count range of longest subsequence
            from_range = []
            count_range = []
            to_range = []

            # Again iterating through the octant list to find the count
            for i in range(len(time)):
                if octant[i] != j:
                    count_check = 0
                else:
                    count_check += 1
                    if count_check == max_count:
                        range_count += 1
                        count_check = 0
                        count_range.append("")
                        # Appending the "From" and "To" time value to the from_range and to_range lists
                        from_range.append(time[i-max_count+1])
                        to_range.append(time[i])

            # Appending the values to respective lists so that the particular column can be printed
            long_sub_len.append(max_count)
            long_sub_len_range.append(max_count)
            octant_ids2.append("Time")
            octant_ids2.extend(count_range)
            long_sub_len_range.append("From")
            long_sub_len_range.extend(from_range)
            count_sub_len.append(range_count)
            count_sub_len_range.append(range_count)
            count_sub_len_range.append("To")
            count_sub_len_range.extend(to_range)

        flag = len(count_sub_len_range)
        # Appending blank spaces
        for i in range(8, len(time)):
            octant_ids.append(" ")
            octant_ids2.append(" ")
            long_sub_len.append("")
            long_sub_len_range.append(" ")
            count_sub_len.append("")
            count_sub_len_range.append(" ")

        # Running a loop to store the values to an output_row list
        for i in range(0, len(time)):
            if i == 0:
                output_row.append([octant_ids[0], long_sub_len[0], count_sub_len[0],
                                  "", octant_ids[0], long_sub_len_range[0], count_sub_len_range[0]])
            else:
                output_row.append([octant_ids[i], long_sub_len[i], count_sub_len[i],
                                  "", octant_ids2[i], long_sub_len_range[i], count_sub_len_range[i]])

        # Writing the values to the output_octant_transition_identify.xlsx file
        # Here i is the range of columns and j is the range of rows. By combinations of characters we are storing the data to the corresponding cells.
        for i in range(19, 26):
            for j in range(4, len(time)):
                cell = 'A'+chr(i+64)+str(j)
                sheet[cell] = output_row[j-4][i-19]

        # Applying borders to the required cells
        for i in range(19, 22):
            for j in range(3, 12):
                cell = 'A'+chr(i+64)+str(j)
                sheet[cell].border = thin_border
        for i in range(23, 26):
            for j in range(3, flag+4):
                cell = 'A'+chr(i+64)+str(j)
                sheet[cell].border = thin_border

    except:
        print("Something went wrong while writing to octant_output.csv")
        st.stop()


# Function - 5 for creating a zip folder


def zipdir(path, ziph):
    # ziph is zipfile handle
    for root, dirs, files in os.walk(path):
        for file in files:
            ziph.write(os.path.join(root, file),
                       os.path.relpath(os.path.join(root, file),
                                       os.path.join(path, '..')))


# Help

# Basic Texts on the Streamlit App
st.title('Octant Batch Processing')
st.header('Project 2')
st.markdown('By Abhinav Mishra (2001MM01) and Hardik Tiwari (2001MM15)')

# Function for the Octant Project GUI


def proj_octant_gui():
    try:

        mode = st.radio("What type of processing do you wish to execute?", ['Single File Processing', 'Multiple Files Processing'], index=0,
                        key=None, help=None, on_change=None, args=None, kwargs=None, disabled=False, horizontal=True, label_visibility="visible")
        uploaded_file=None
        multiplefilespath=''
        if mode == 'Single File Processing':
            # Field to upload the file
            uploaded_file = st.file_uploader(label="Input File", type=[
                'xlsx'], accept_multiple_files=False)
            try:
                # Mod Value taken as input from the USER (Default Value is 5000)
                mod = int(st.text_input(label="Mod Value",
                                        placeholder="Enter a mod value for computation", value="5000"))
            except ValueError:
                st.write("Please enter Valid Input in the Mod Value field")
                st.stop()

        elif mode == 'Multiple Files Processing':
            # Path of the folder, If the USER wants multiple files to be computed
            multiplefilespath = st.text_input(label="Enter the Folder path if you want to compute multiple files:",
                                              placeholder="Folder Path")
            try:
            # Mod Value taken as input from the USER (Default Value is 5000)
                mod = int(st.text_input(label="Mod Value",
                                    placeholder="Enter a mod value for computation", value="5000"))
            except ValueError:
                st.write("Please enter Valid Input in the Mod Value field")
                st.stop()

        # Compute button to carry out the computation
        compute = st.button(label="Compute")

        # Parent Directory
        parent = ""

        # If the Output directory is not present, make an output directory called octant_analysis_output in Downloads folder of the USER
        if (os.path.exists(os.path.expanduser('~').replace('\\', '/')+'/Downloads/octant_analysis_output/') == False):
            os.makedirs(os.path.expanduser('~').replace(
                '\\', '/')+'/Downloads/octant_analysis_output/')
            parent = os.path.expanduser('~').replace(
                '\\', '/')+'/Downloads/octant_analysis_output/'

        # Else If the Output directory is present
        else:

            # Delete the the Output Directory
            shutil.rmtree((os.path.expanduser('~').replace(
                '\\', '/')+'/Downloads/octant_analysis_output/'))

            # Make a new Output Directory
            os.makedirs(os.path.expanduser('~').replace(
                '\\', '/')+'/Downloads/octant_analysis_output/')
            parent = os.path.expanduser('~').replace(
                '\\', '/')+'/Downloads/octant_analysis_output/'

        # If the compute button is clicked
        if (compute):
            # If a single file is not uploaded and the USER uploads the path of the directory containing the files.
            if uploaded_file == None and multiplefilespath != '':

                # Changing the current directory according to the USER's path
                os.chdir(multiplefilespath)

                # Iterating through the files in the USER's directory
                count=0
                for inputfile in os.listdir():
                    if os.path.isfile(inputfile):
                        count=1
                        if inputfile.endswith(".xls") or inputfile.endswith(".xlsx"):
                            wb = Workbook()
                            sheet = wb.active

                            # Calling the respective functions
                            octant_range_names(sheet, inputfile, mod)
                            octant_transition_count(sheet, inputfile, mod)
                            octant_longest_subsequence_count_with_range(
                                sheet, inputfile)

                            # Adjusting Column Widths
                            for col in sheet.columns:
                                max_length = 0
                                # Get the column name
                                column = col[0].column_letter
                                for cell in col:
                                    try:  # Necessary to avoid error on empty cells
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = (max_length+1)
                                sheet.column_dimensions[column].width = adjusted_width

                            # Name of the input file
                            inputfile = inputfile[:-5]

                            # Giving the Name Format
                            datetimenow = str(datetime.now())
                            datetimeformat = "_" + \
                                datetimenow[:10]+"-"+datetimenow[11:13]+"-" + \
                                datetimenow[14:16]+"-"+datetimenow[17:19]

                            # The workbook is saved in the downloads' output folder of the USER
                            wb.save(parent+str(inputfile)+"_" +
                                    str(mod)+datetimeformat+".xlsx")
                            wb.close()  # Closing the workbook.

                            # Output a message that the Outputfile is generated for the given input file
                            st.write("OutputFile Generated for:", inputfile)
                        else:
                            continue
                    else:
                        continue
                if count==0:
                    st.write("No Excel File(s) exists in the specified path!")
                    st.stop()

            # If both the fields are not filled, a message shows up to the USER
            elif uploaded_file == None and multiplefilespath == '':
                st.markdown("Please enter the path or upload a single file!")
                st.stop()

            # If a single file is uploaded
            else:
                inputfile = uploaded_file
                wb = Workbook()
                sheet = wb.active

                # Calling the respective functions
                octant_range_names(sheet, inputfile, mod)
                octant_transition_count(sheet, inputfile, mod)
                octant_longest_subsequence_count_with_range(
                    sheet, inputfile)

                # Adjusting Column Widths
                for col in sheet.columns:
                    max_length = 0
                    # Get the column name
                    column = col[0].column_letter
                    for cell in col:
                        try:  # Necessary to avoid error on empty cells
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length+1)
                    sheet.column_dimensions[column].width = adjusted_width

                # Name of the input file
                inputfilenameformat = inputfile.name
                inputfile.name = inputfile.name[:-5]

                # Giving the Name Format
                datetimenow = str(datetime.now())
                datetimeformat = "_" + \
                    datetimenow[:10]+"-"+datetimenow[11:13]+"-" + \
                    datetimenow[14:16]+"-"+datetimenow[17:19]

                # The workbook is saved in the downloads' output folder of the USER
                wb.save(parent+str(inputfile.name)+"_" +
                        str(mod)+datetimeformat+".xlsx")

                wb.close()  # Closing the workbook.

                # Output a message that the Outputfile is generated for the given input file
                st.write("OutputFile Generated for:",
                            inputfilenameformat)


            # Making a zip folder of the Output Folder
            shutil.make_archive(parent, 'zip', parent)

            # Telling the USER that the files are downloaded in their Downloads folder, Also giving them an option to download the same
            st.caption("The Files are downloaded and are in your Downloads folder. If You cannot find them you can still download them by Clicking on 'Download ZIP' Button")
            with open(os.path.dirname(os.path.dirname(parent))+"/octant_analysis_output.zip", "rb") as fp:
                st.download_button(
                    label="Download ZIP",
                    data=fp,
                    file_name="Output.zip",
                    mime="application/zip"
                )
            fp.close()

        # Deleting the Output folder as a zip file is already generated for the same
        shutil.rmtree((os.path.expanduser('~').replace(
            '\\', '/')+'/Downloads/octant_analysis_output/'))

    # If the path entered is invalid
    except FileNotFoundError:
        st.markdown("The Path Entered is invalid!")


# Calling Main Function of the program
proj_octant_gui()

# This shall be the last lines of the code.
end_time = datetime.now()

# Deleting the Output folder as a zip file is already generated for the same
try:
    os.rmdir(os.path.expanduser('~').replace(
        '\\', '/')+'/Downloads/octant_analysis_output/')
except:
    pass
